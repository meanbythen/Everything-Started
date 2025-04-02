import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from get_coordinates import process_pdf
import threading
import queue
import json
import traceback
import logging

# 设置日志记录
def setup_logging():
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'invoice_processor.log')
    
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return log_file

def show_error_and_wait(error_msg):
    """显示错误信息并等待用户确认"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    messagebox.showerror("错误", f"程序启动失败:\n{error_msg}\n\n详细日志已保存到logs文件夹")
    
def resource_path(relative_path):
    """获取资源的绝对路径，用于处理打包后的路径"""
    try:
        # PyInstaller创建临时文件夹，将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 禁用 pdf2image 的警告
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

class InvoiceProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("发票处理程序")
        self.root.geometry("800x600")  # 设置窗口大小
        
        # 捕获窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 设置窗口图标（如果有的话）
        try:
            icon_path = resource_path("icon.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"加载图标失败: {str(e)}")
            
        self.setup_ui()
        self.processing_queue = queue.Queue()
        logging.info("GUI初始化完成")
        
    def on_closing(self):
        """处理窗口关闭事件"""
        if messagebox.askokcancel("确认", "确定要退出程序吗？"):
            logging.info("用户关闭程序")
            self.root.destroy()

    def setup_ui(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 源文件选择
        ttk.Label(main_frame, text="源文件:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.source_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.source_path, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="浏览", command=self.browse_source).grid(row=0, column=2, padx=5)

        # 输出文件夹选择
        ttk.Label(main_frame, text="输出文件夹:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.output_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.output_path, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(main_frame, text="浏览", command=self.browse_output).grid(row=1, column=2, padx=5)

        # 处理模式选择
        ttk.Label(main_frame, text="处理模式:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.process_mode = tk.StringVar(value="folder")
        ttk.Radiobutton(main_frame, text="处理文件夹", variable=self.process_mode, 
                      value="folder").grid(row=2, column=1, sticky="w")
        ttk.Radiobutton(main_frame, text="处理单个文件", variable=self.process_mode, 
                      value="single").grid(row=2, column=1)

        # 文件列表
        ttk.Label(main_frame, text="文件列表:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.file_listbox = tk.Listbox(main_frame, width=60, height=10)
        self.file_listbox.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

        # 添加滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.file_listbox.yview)
        scrollbar.grid(row=4, column=3, sticky="ns")
        self.file_listbox.configure(yscrollcommand=scrollbar.set)

        # 开始处理按钮
        ttk.Button(main_frame, text="开始处理", command=self.start_processing).grid(row=5, column=1, pady=10)

        # 处理进度标签
        self.progress_label = ttk.Label(main_frame, text="处理进度")
        self.progress_label.grid(row=6, column=0, columnspan=3, pady=5)

        # 配置网格权重
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(4, weight=1)

    def browse_source(self):
        if self.process_mode.get() == "folder":
            path = filedialog.askdirectory()
        else:
            path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        
        if path:
            self.source_path.set(path)
            self.update_file_list()

    def browse_output(self):
        path = filedialog.askdirectory()
        if path:
            self.output_path.set(path)

    def update_file_list(self):
        self.file_listbox.delete(0, tk.END)
        path = self.source_path.get()
        
        if self.process_mode.get() == "folder":
            if os.path.isdir(path):
                for file in os.listdir(path):
                    if file.lower().endswith('.pdf'):
                        self.file_listbox.insert(tk.END, file)
        else:
            if os.path.isfile(path) and path.lower().endswith('.pdf'):
                self.file_listbox.insert(tk.END, os.path.basename(path))

    def process_files(self):
        source_path = self.source_path.get()
        output_path = self.output_path.get()
        
        if not source_path or not output_path:
            messagebox.showerror("错误", "请选择源文件/文件夹和输出文件夹")
            return

        # 准备存储所有发票数据的列表
        all_invoice_data = []
        temp_files = []  # 用于记录需要清理的临时文件
        
        try:
            # 获取要处理的PDF文件列表
            pdf_files = []
            if self.process_mode.get() == "folder":
                if os.path.isdir(source_path):
                    pdf_files = [f for f in os.listdir(source_path) if f.lower().endswith('.pdf')]
                    pdf_files = [os.path.join(source_path, f) for f in pdf_files]
            else:
                if os.path.isfile(source_path) and source_path.lower().endswith('.pdf'):
                    pdf_files = [source_path]
            
            if not pdf_files:
                messagebox.showwarning("警告", "没有找到PDF文件")
                return
                
            total_files = len(pdf_files)
            
            # 处理每个PDF文件
            for index, pdf_path in enumerate(pdf_files, 1):
                pdf_file = os.path.basename(pdf_path)
                self.progress_label.config(text=f"正在处理PDF: {pdf_file} ({index}/{total_files})")
                self.root.update()
                
                try:
                    # 处理PDF文件生成JSON
                    process_pdf(pdf_path)
                    # 记录生成的临时文件
                    base_name = os.path.splitext(pdf_file)[0]
                    json_path = os.path.join(os.path.dirname(pdf_path), f"{base_name}.json")
                    csv_path = os.path.join(os.path.dirname(pdf_path), f"{base_name}.csv")
                    temp_files.extend([json_path, csv_path])
                    
                    # 读取生成的JSON文件
                    try:
                        with open(json_path, 'r', encoding='utf-8') as f:
                            data = json.loads(f.read())
                            # 提取所需字段
                            invoice_data = {
                                "文件名": pdf_file,
                                "发票类型": data.get("invoice_type", ""),
                                "发票号码": data.get("invoice_number", ""),
                                "开票日期": data.get("invoice_date", ""),
                                "采购方名称": data.get("buyer_name", ""),
                                "采购方纳税人识别号": data.get("buyer_tax_id", ""),
                                "销售方名称": data.get("seller_name", ""),
                                "销售方纳税人识别号": data.get("seller_tax_id", ""),
                                "金额": data.get("net_amount", ""),
                                "税额": data.get("tax_amount", ""),
                                "价税合计": data.get("total_amount", "")
                            }
                            all_invoice_data.append(invoice_data)
                            print(f"成功处理文件: {pdf_file}")
                    except Exception as e:
                        print(f"读取JSON文件失败: {str(e)}")
                        messagebox.showwarning("警告", f"处理文件 {pdf_file} 时出错: {str(e)}")
                        continue
                        
                except Exception as e:
                    print(f"处理PDF文件 {pdf_file} 时出错: {str(e)}")
                    messagebox.showwarning("警告", f"处理文件 {pdf_file} 时出错: {str(e)}")
                    continue

            # 使用openpyxl处理Excel文件
            if all_invoice_data:
                try:
                    excel_path = os.path.join(output_path, "发票数据汇总.xlsx")
                    os.makedirs(output_path, exist_ok=True)
                    
                    # 定义列标题
                    columns = ["文件名", "发票类型", "发票号码", "开票日期", 
                             "采购方名称", "采购方纳税人识别号", 
                             "销售方名称", "销售方纳税人识别号", 
                             "金额", "税额", "价税合计"]
                    
                    # 检查是否已存在Excel文件
                    existing_invoice_numbers = set()
                    if os.path.exists(excel_path):
                        try:
                            # 尝试加载现有文件
                            wb = load_workbook(excel_path)
                            ws = wb.active
                            
                            # 验证表头是否匹配
                            existing_headers = [cell.value for cell in ws[1]]
                            if existing_headers != columns:
                                raise ValueError("现有Excel文件的表头与程序不匹配")
                            
                            # 获取现有发票号码
                            invoice_number_col = columns.index("发票号码") + 1
                            for row in range(2, ws.max_row + 1):
                                invoice_number = ws.cell(row=row, column=invoice_number_col).value
                                if invoice_number:
                                    existing_invoice_numbers.add(invoice_number)
                                    
                        except Exception as e:
                            error_msg = str(e)
                            message = '读取现有Excel文件时出错: {}\n是否要创建新文件？\n(选择"是"将备份原文件并创建新文件，选择"否"将取消操作)'.format(error_msg)
                            user_choice = messagebox.askyesno("错误", message)
                            if user_choice:
                                # 如果文件存在，先备份
                                if os.path.exists(excel_path):
                                    backup_path = excel_path + '.bak'
                                    try:
                                        os.rename(excel_path, backup_path)
                                        print(f"已将原文件备份为: {backup_path}")
                                    except Exception as e:
                                        print(f"备份文件时出错: {str(e)}")
                                # 创建新的工作簿
                                wb = Workbook()
                                ws = wb.active
                                # 写入表头
                                for col, header in enumerate(columns, 1):
                                    ws.cell(row=1, column=col, value=header)
                    else:
                        # 如果文件不存在，创建新的
                        wb = Workbook()
                        ws = wb.active
                        # 写入表头
                        for col, header in enumerate(columns, 1):
                            ws.cell(row=1, column=col, value=header)
                    
                    # 添加新数据
                    duplicate_invoices = []
                    for invoice_data in all_invoice_data:
                        invoice_number = invoice_data["发票号码"]
                        if invoice_number in existing_invoice_numbers:
                            duplicate_invoices.append(invoice_number)
                            continue
                        
                        # 添加新行
                        row_data = [invoice_data[col] for col in columns]
                        ws.append(row_data)
                        existing_invoice_numbers.add(invoice_number)
                    
                    # 保存Excel文件
                    wb.save(excel_path)
                    print(f"数据已保存到: {excel_path}")
                    
                    # 显示处理结果
                    success_message = f"处理完成！\n成功处理 {len(all_invoice_data)} 个文件"
                    if duplicate_invoices:
                        success_message += f"\n发现 {len(duplicate_invoices)} 个重复发票号码，已跳过"
                    messagebox.showinfo("完成", success_message)
                    
                except Exception as e:
                    print(f"保存Excel文件时出错: {str(e)}")
                    messagebox.showerror("错误", f"保存Excel文件时出错: {str(e)}")
        
        except Exception as e:
            print(f"处理过程中出错: {str(e)}")
            messagebox.showerror("错误", f"处理过程中出错: {str(e)}")
        
        finally:
            # 清理临时文件
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print(f"已删除临时文件: {temp_file}")
                except Exception as e:
                    print(f"删除临时文件 {temp_file} 时出错: {str(e)}")
            
            self.progress_label.config(text="就绪")

    def start_processing(self):
        threading.Thread(target=self.process_files, daemon=True).start()

# 移除直接执行代码，因为我们使用启动器
if __name__ == '__main__':
    print("请使用 run.py 启动程序") 